import numpy as np
import matplotlib.pyplot as plt
from matplotlib import font_manager
from matplotlib.patches import Rectangle
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================
# matplotlib 中文字体设置，解决中文乱码
# =============================
def setup_chinese_font():
    candidate_fonts = [
        "Microsoft YaHei",
        "SimHei",
        "SimSun",
        "NSimSun",
        "KaiTi",
        "FangSong",
        "DengXian",
        "Noto Sans CJK SC",
        "Source Han Sans SC",
        "WenQuanYi Zen Hei"
    ]

    available_fonts = {f.name for f in font_manager.fontManager.ttflist}

    for font_name in candidate_fonts:
        if font_name in available_fonts:
            plt.rcParams["font.sans-serif"] = [font_name]
            plt.rcParams["axes.unicode_minus"] = False
            print(f"已启用中文字体: {font_name}")
            return

    plt.rcParams["font.sans-serif"] = ["DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
    print("警告：未找到合适的中文字体，图中中文仍可能乱码，请安装微软雅黑或黑体。")


setup_chinese_font()


# =============================
# 原始数据
# =============================
distance = np.array([
    0.0000, 0.8331, 3.6807, 13.8020, 25.1790, 35.4950, 46.0177, 57.0600,
    67.3924, 84.3925, 101.9415, 121.0518, 140.7804, 160.9664, 181.6260,
    202.9510, 225.0580, 247.0480, 271.2910, 295.3940, 320.1240, 344.0730,
    358.6020, 387.3670, 421.0620, 438.6900, 453.2400, 469.3380, 485.6700, 500.0000
])

temperature = np.array([
    649.302, 648.285, 646.789, 615.4, 587.037, 563.738, 545.841, 527.104,
    513.387, 493.408, 477.14, 464.483, 456.835, 448.064, 439.41,
    432.917, 425.066, 416.81, 408.695, 399.782, 389.785, 380.018,
    374.458, 369.644, 358.265, 353.435, 349.444, 347.157, 348.54, 354.843
])


# =============================
# 参数设置
# =============================
TOTAL_LENGTH = 500.0
MAX_ZONES = 8
ALPHA = 5.0
MIN_MODULES_PER_ZONE = 1
EQUAL_K = 8

MODULE_LENGTH = 23.0          # 单个灯管模块本体长度(mm)
MODULE_GAP = 10.0             # 相邻模块间距(mm)
MODULE_PITCH = MODULE_LENGTH + MODULE_GAP

OUTER_EDGE_ALLOW = 10.0       # 左右最外边缘允许外超长度(mm)

MIN_SIZE = MODULE_LENGTH
EXPORT_FILENAME = "分区结果_灯管间距10mm_含模块实际排布_完整版.xlsx"


# =============================
# 基础函数
# =============================
def calc_install_length(modules, module_length=23.0, module_gap=10.0):
    """
    N个模块总安装长度 = N*模块长度 + (N-1)*间距
    """
    if modules <= 0:
        return 0.0
    return modules * module_length + (modules - 1) * module_gap


def calc_gap_length(modules, module_gap=10.0):
    if modules <= 1:
        return 0.0
    return (modules - 1) * module_gap


def calc_max_fittable_modules(zone_size, is_first_zone=False, is_last_zone=False,
                              module_length=23.0, module_gap=10.0, outer_edge_allow=10.0):
    """
    计算在给定分区中最多能放下多少个模块

    规则：
    1. 内部分区：安装长度 <= 分区长度
    2. 首分区：允许向左边界外超出 outer_edge_allow
    3. 末分区：允许向右边界外超出 outer_edge_allow
    4. 不允许0模块，最少返回1
    """
    effective_size = zone_size
    if is_first_zone:
        effective_size += outer_edge_allow
    if is_last_zone:
        effective_size += outer_edge_allow

    if effective_size < module_length:
        return 1

    n = math.floor((effective_size + module_gap) / (module_length + module_gap))
    return max(1, n)


def compute_actual_overhang(zone_size, modules, is_first_zone=False, is_last_zone=False,
                            module_length=23.0, module_gap=10.0):
    """
    计算实际安装长度，以及左右外超长度
    """
    install_length = calc_install_length(modules, module_length, module_gap)
    exceed = max(0.0, install_length - zone_size)

    left_overhang = 0.0
    right_overhang = 0.0

    if exceed > 0:
        if is_first_zone and not is_last_zone:
            left_overhang = exceed
        elif is_last_zone and not is_first_zone:
            right_overhang = exceed
        elif is_first_zone and is_last_zone:
            left_overhang = exceed / 2.0
            right_overhang = exceed / 2.0

    return install_length, exceed, left_overhang, right_overhang


def compute_module_positions_in_zone(x0, x1, modules,
                                     is_first=False, is_last=False,
                                     module_length=23.0, module_gap=10.0):
    """
    计算某个分区内每个模块的实际起止位置 [start, end]

    规则：
    1. 内部分区：模块串完全放在分区内部，并居中排布
    2. 首分区：若安装长度超过分区长度，则只允许向左外超
    3. 末分区：若安装长度超过分区长度，则只允许向右外超
    """
    zone_length = x1 - x0
    install_length = calc_install_length(modules, module_length, module_gap)

    if is_first and install_length > zone_length:
        start_pos = x1 - install_length
    elif is_last and install_length > zone_length:
        start_pos = x0
    else:
        start_pos = x0 + 0.5 * (zone_length - install_length)

    positions = []
    for i in range(modules):
        xs = start_pos + i * (module_length + module_gap)
        xe = xs + module_length
        positions.append((xs, xe))

    return positions


# =============================
# 计算梯度权重
# =============================
def compute_weights(dist, temp, alpha=5.0):
    grad = np.zeros(len(temp))
    for i in range(1, len(temp)):
        dx = dist[i] - dist[i - 1]
        if dx > 0:
            grad[i] = abs(temp[i] - temp[i - 1]) / dx

    gmax = np.max(grad) if np.max(grad) > 0 else 1.0
    grad = grad / gmax
    return 1.0 + alpha * grad


# =============================
# 提取区间数据（含边界插值）
# =============================
def get_segment_points_with_interpolation(dist, temp, x0, x1):
    mask = (dist > x0) & (dist < x1)
    inner_x = dist[mask]
    inner_t = temp[mask]

    t0 = np.interp(x0, dist, temp)
    t1 = np.interp(x1, dist, temp)

    x_all = np.concatenate(([x0], inner_x, [x1]))
    t_all = np.concatenate(([t0], inner_t, [t1]))

    order = np.argsort(x_all)
    x_all = x_all[order]
    t_all = t_all[order]

    return x_all, t_all


# =============================
# 连续等距分区
# =============================
def build_equal_edges(total_length, k):
    return np.linspace(0.0, total_length, k + 1)


def equal_partition_continuous(total_length, k):
    edges = build_equal_edges(total_length, k)
    zones = []
    for i in range(k):
        zones.append((edges[i], edges[i + 1]))
    return zones


def build_zones_equal_continuous(dist, temp, zones_equal):
    zones = []
    total_zones = len(zones_equal)

    for idx, (x0, x1) in enumerate(zones_equal):
        seg_x, seg_t = get_segment_points_with_interpolation(dist, temp, x0, x1)
        size = x1 - x0
        is_first = (idx == 0)
        is_last = (idx == total_zones - 1)

        modules = calc_max_fittable_modules(
            size,
            is_first_zone=is_first,
            is_last_zone=is_last,
            module_length=MODULE_LENGTH,
            module_gap=MODULE_GAP,
            outer_edge_allow=OUTER_EDGE_ALLOW
        )

        avg_temp = np.mean(seg_t)
        install_length, exceed, left_overhang, right_overhang = compute_actual_overhang(
            size, modules,
            is_first_zone=is_first,
            is_last_zone=is_last,
            module_length=MODULE_LENGTH,
            module_gap=MODULE_GAP
        )

        body_length = modules * MODULE_LENGTH
        gap_length = calc_gap_length(modules, MODULE_GAP)

        zones.append({
            "id": idx + 1,
            "range": (x0, x1),
            "size": size,
            "modules": modules,
            "avg_temp": avg_temp,
            "segment_x": seg_x,
            "segment_t": seg_t,
            "install_length": install_length,
            "body_length": body_length,
            "gap_length": gap_length,
            "exceed_length": exceed,
            "left_overhang": left_overhang,
            "right_overhang": right_overhang,
            "is_first": is_first,
            "is_last": is_last
        })
    return zones


# =============================
# 严格模块对齐分区
# 分区边界按安装节距离散
# =============================
def build_module_edges(total_length=500.0, pitch=33.0):
    n_full = int(total_length // pitch)
    edges = [i * pitch for i in range(n_full + 1)]
    if edges[-1] < total_length:
        edges.append(total_length)
    return np.array(edges)


def segment_cost_on_edges(dist, temp, weights, x0, x1):
    seg_x, seg_t = get_segment_points_with_interpolation(dist, temp, x0, x1)
    if len(seg_t) < 2:
        return np.inf, None

    seg_w = np.interp(seg_x, dist, weights)
    mean_t = np.sum(seg_w * seg_t) / np.sum(seg_w)
    cost = np.sum(seg_w * (seg_t - mean_t) ** 2)
    return cost, mean_t


def optimal_partition_aligned(dist, temp, weights):
    edges = build_module_edges(TOTAL_LENGTH, MODULE_PITCH)
    m = len(edges)

    dp = np.full((MAX_ZONES + 1, m), np.inf)
    split = np.full((MAX_ZONES + 1, m), -1, dtype=int)

    dp[0, 0] = 0.0

    cost_table = [[np.inf] * m for _ in range(m)]
    mean_table = [[None] * m for _ in range(m)]

    for a in range(m - 1):
        for b in range(a + 1, m):
            x0, x1 = edges[a], edges[b]
            size = x1 - x0

            if size < MIN_SIZE:
                continue

            cost, mean_t = segment_cost_on_edges(dist, temp, weights, x0, x1)
            cost_table[a][b] = cost
            mean_table[a][b] = mean_t

    for k in range(1, MAX_ZONES + 1):
        for j in range(1, m):
            for i in range(j):
                if np.isinf(cost_table[i][j]):
                    continue
                total = dp[k - 1, i] + cost_table[i][j]
                if total < dp[k, j]:
                    dp[k, j] = total
                    split[k, j] = i

    end_idx = m - 1
    best_k = np.argmin(dp[1:, end_idx]) + 1

    zones_idx = []
    j = end_idx
    k = best_k
    while k > 0:
        i = split[k, j]
        zones_idx.append((i, j))
        j = i
        k -= 1

    zones_idx.reverse()
    return edges, zones_idx, mean_table


def build_zones_aligned(edges, zones_idx, mean_table):
    zones = []
    total_zones = len(zones_idx)

    for idx, (a, b) in enumerate(zones_idx):
        x0, x1 = edges[a], edges[b]
        size = x1 - x0
        is_first = (idx == 0)
        is_last = (idx == total_zones - 1)

        modules = calc_max_fittable_modules(
            size,
            is_first_zone=is_first,
            is_last_zone=is_last,
            module_length=MODULE_LENGTH,
            module_gap=MODULE_GAP,
            outer_edge_allow=OUTER_EDGE_ALLOW
        )

        install_length, exceed, left_overhang, right_overhang = compute_actual_overhang(
            size, modules,
            is_first_zone=is_first,
            is_last_zone=is_last,
            module_length=MODULE_LENGTH,
            module_gap=MODULE_GAP
        )

        body_length = modules * MODULE_LENGTH
        gap_length = calc_gap_length(modules, MODULE_GAP)

        zones.append({
            "id": idx + 1,
            "range": (x0, x1),
            "size": size,
            "modules": modules,
            "avg_temp": mean_table[a][b],
            "install_length": install_length,
            "body_length": body_length,
            "gap_length": gap_length,
            "exceed_length": exceed,
            "left_overhang": left_overhang,
            "right_overhang": right_overhang,
            "is_first": is_first,
            "is_last": is_last
        })
    return zones


# =============================
# 统一评价函数
# =============================
def evaluate_zoning_quality(dist, temp, zones):
    N = len(temp)
    K = len(zones)

    zone_means = []
    zone_sizes = []
    fit_error = 0.0
    heater_errors = []
    internal_violation_count = 0

    for z in zones:
        x0, x1 = z["range"]
        seg_x, seg_t = get_segment_points_with_interpolation(dist, temp, x0, x1)
        mean = np.mean(seg_t)

        zone_means.append(mean)
        zone_sizes.append(z["size"])
        fit_error += np.sum((seg_t - mean) ** 2)

        install_length = z["install_length"]
        size = z["size"]

        if z["is_first"] or z["is_last"]:
            effective_size = size + OUTER_EDGE_ALLOW
            if install_length > effective_size + 1e-9:
                internal_violation_count += 1
            heater_errors.append(abs(size - install_length))
        else:
            if install_length > size + 1e-9:
                internal_violation_count += 1
            heater_errors.append(abs(size - install_length))

    E_fit = fit_error / N if N > 0 else np.inf
    E_sep = np.mean([abs(zone_means[i + 1] - zone_means[i]) for i in range(K - 1)]) if K > 1 else 0.0
    size_compliance = sum(1 for s in zone_sizes if s >= MIN_SIZE) / K if K > 0 else 0.0
    E_heater = np.mean(heater_errors) if len(heater_errors) > 0 else np.inf
    cv = np.std(zone_sizes) / np.mean(zone_sizes) if K > 1 else 0.0

    S_fit = np.exp(-E_fit / 1000.0)
    S_sep = E_sep / (max(zone_means) - min(zone_means) + 1e-6) if len(zone_means) > 1 else 0.0
    S_heater = np.exp(-E_heater / 10.0)
    S_balance = 1.0 / (1.0 + cv)

    violation_penalty = 0.15 * internal_violation_count

    score = 0.35 * S_fit + 0.25 * S_sep + 0.2 * S_heater + 0.2 * S_balance - violation_penalty
    score = max(0.0, score)

    total_modules = sum(z["modules"] for z in zones)
    total_install_length = sum(z["install_length"] for z in zones)
    total_body_length = sum(z["body_length"] for z in zones)
    total_gap_length = sum(z["gap_length"] for z in zones)
    total_left_overhang = sum(z["left_overhang"] for z in zones)
    total_right_overhang = sum(z["right_overhang"] for z in zones)

    return {
        "E_fit": E_fit,
        "E_sep": E_sep,
        "尺寸合规率": size_compliance,
        "安装长度误差": E_heater,
        "均衡性": S_balance,
        "内部违规数": internal_violation_count,
        "综合得分": score,
        "总模块数": total_modules,
        "总安装占位长度": total_install_length,
        "总模块本体长度": total_body_length,
        "总间距长度": total_gap_length,
        "左侧总外超长度": total_left_overhang,
        "右侧总外超长度": total_right_overhang
    }


def print_metrics(title, metrics):
    print(f"\n=== {title} ===")
    for k, v in metrics.items():
        if isinstance(v, float):
            print(f"{k}: {v:.4f}")
        else:
            print(f"{k}: {v}")


# =============================
# 绘图辅助：在分区上画模块真实位置
# =============================
def draw_modules_on_zone(ax, zone, color_fill, color_edge,
                         module_height=6.0,
                         module_length=23.0, module_gap=10.0):
    x0, x1 = zone["range"]
    y = zone["avg_temp"]
    modules = zone["modules"]
    is_first = zone.get("is_first", False)
    is_last = zone.get("is_last", False)

    module_positions = compute_module_positions_in_zone(
        x0, x1, modules,
        is_first=is_first,
        is_last=is_last,
        module_length=module_length,
        module_gap=module_gap
    )

    for idx, (xs, xe) in enumerate(module_positions, start=1):
        rect = Rectangle(
            (xs, y - module_height / 2.0),
            xe - xs,
            module_height,
            facecolor=color_fill,
            edgecolor=color_edge,
            linewidth=1.2,
            alpha=0.85
        )
        ax.add_patch(rect)

        ax.text(
            0.5 * (xs + xe),
            y,
            f"{idx}",
            ha="center",
            va="center",
            fontsize=7,
            color="black"
        )


# =============================
# 绘图：分区对比图（含模块真实排布）
# =============================
def plot_compare_equal_vs_aligned(dist, temp, zones_equal, zones_aligned, edges, metrics_equal, metrics_aligned):
    fig, axes = plt.subplots(2, 1, figsize=(14, 11), sharex=True)

    # 上图：连续等距分区
    ax = axes[0]
    ax.plot(dist, temp, 'o-', color='black', linewidth=1.2, markersize=4, label="原始温度曲线")

    for z in zones_equal:
        x0, x1 = z["range"]
        y = z["avg_temp"]
        xm = 0.5 * (x0 + x1)

        ax.hlines(y, x0, x1, linewidth=3.2, color='tab:blue', alpha=0.95)
        ax.axvline(x0, linestyle='--', alpha=0.25, color='gray')

        draw_modules_on_zone(
            ax, z,
            color_fill="#9ecae1",
            color_edge="tab:blue",
            module_height=7.0,
            module_length=MODULE_LENGTH,
            module_gap=MODULE_GAP
        )

        extra_text = ""
        if z["left_overhang"] > 0:
            extra_text += f"\n左外超={z['left_overhang']:.1f} mm"
        if z["right_overhang"] > 0:
            extra_text += f"\n右外超={z['right_overhang']:.1f} mm"

        ax.text(
            xm, y + 14,
            f'Z{z["id"]}\n'
            f'{z["modules"]}个模块\n'
            f'安装长={z["install_length"]:.1f} mm\n'
            f'{x0:.1f}-{x1:.1f} mm{extra_text}',
            ha='center', va='bottom', fontsize=9,
            bbox=dict(boxstyle='round,pad=0.25', fc='white', ec='tab:blue', alpha=0.9)
        )

    ax.axvline(TOTAL_LENGTH, linestyle='--', alpha=0.4, color='black')
    ax.set_title(
        f'连续等距分区 | 得分={metrics_equal["综合得分"]:.3f}, '
        f'E_fit={metrics_equal["E_fit"]:.2f}, E_sep={metrics_equal["E_sep"]:.2f}'
    )
    ax.set_ylabel("温度 (℃)")
    ax.grid(True, alpha=0.3)

    # 下图：严格模块对齐分区
    ax = axes[1]
    ax.plot(dist, temp, 'o-', color='black', linewidth=1.2, markersize=4, label="原始温度曲线")

    for e in edges:
        ax.axvline(e, color='lightgray', linestyle=':', alpha=0.35)

    for z in zones_aligned:
        x0, x1 = z["range"]
        y = z["avg_temp"]
        xm = 0.5 * (x0 + x1)

        ax.hlines(y, x0, x1, linewidth=3.2, color='tab:red', alpha=0.95)
        ax.axvline(x0, linestyle='--', alpha=0.25, color='gray')

        draw_modules_on_zone(
            ax, z,
            color_fill="#fcbba1",
            color_edge="tab:red",
            module_height=7.0,
            module_length=MODULE_LENGTH,
            module_gap=MODULE_GAP
        )

        extra_text = ""
        if z["left_overhang"] > 0:
            extra_text += f"\n左外超={z['left_overhang']:.1f} mm"
        if z["right_overhang"] > 0:
            extra_text += f"\n右外超={z['right_overhang']:.1f} mm"

        ax.text(
            xm, y + 14,
            f'Z{z["id"]}\n'
            f'{z["modules"]}个模块\n'
            f'安装长={z["install_length"]:.1f} mm\n'
            f'{x0:.1f}-{x1:.1f} mm{extra_text}',
            ha='center', va='bottom', fontsize=9,
            bbox=dict(boxstyle='round,pad=0.25', fc='white', ec='tab:red', alpha=0.9)
        )

    ax.axvline(TOTAL_LENGTH, linestyle='--', alpha=0.4, color='black')
    ax.set_title(
        f'严格模块对齐分区 | 得分={metrics_aligned["综合得分"]:.3f}, '
        f'E_fit={metrics_aligned["E_fit"]:.2f}, E_sep={metrics_aligned["E_sep"]:.2f}'
    )
    ax.set_xlabel("距离 (mm)")
    ax.set_ylabel("温度 (℃)")
    ax.grid(True, alpha=0.3)

    plt.suptitle("连续等距分区 与 严格模块对齐分区 对比（含模块实际排布位置）", fontsize=15)
    plt.tight_layout()
    plt.show()


# =============================
# 绘图：独立安装排布图
# =============================
def plot_module_layout(zones_equal, zones_aligned):
    fig, axes = plt.subplots(2, 1, figsize=(14, 8), sharex=True)

    def draw_layout(ax, zones, title, face_color, edge_color):
        row_height = 0.6
        total = len(zones)

        for idx, z in enumerate(zones, start=1):
            y = total - idx + 1
            x0, x1 = z["range"]

            # 分区背景
            zone_rect = Rectangle(
                (x0, y - row_height / 2),
                x1 - x0,
                row_height,
                facecolor="none",
                edgecolor="gray",
                linestyle="--",
                linewidth=1.0,
                alpha=0.8
            )
            ax.add_patch(zone_rect)

            # 模块位置
            positions = compute_module_positions_in_zone(
                x0, x1, z["modules"],
                is_first=z["is_first"],
                is_last=z["is_last"],
                module_length=MODULE_LENGTH,
                module_gap=MODULE_GAP
            )

            for m_idx, (xs, xe) in enumerate(positions, start=1):
                rect = Rectangle(
                    (xs, y - 0.18),
                    xe - xs,
                    0.36,
                    facecolor=face_color,
                    edgecolor=edge_color,
                    linewidth=1.2,
                    alpha=0.9
                )
                ax.add_patch(rect)

                ax.text(
                    0.5 * (xs + xe), y,
                    str(m_idx),
                    ha='center', va='center', fontsize=8
                )

            label = f'Z{z["id"]} | {x0:.1f}-{x1:.1f} mm | {z["modules"]}模块'
            if z["left_overhang"] > 0:
                label += f' | 左外超={z["left_overhang"]:.1f}'
            if z["right_overhang"] > 0:
                label += f' | 右外超={z["right_overhang"]:.1f}'

            ax.text(-25, y, label, ha='left', va='center', fontsize=9)

        ax.axvline(0, color='black', linewidth=1.0, alpha=0.7)
        ax.axvline(TOTAL_LENGTH, color='black', linewidth=1.0, alpha=0.7)
        ax.set_ylim(0.5, total + 0.8)
        ax.set_yticks(range(1, total + 1))
        ax.set_yticklabels([f"第{total-i+1}行" for i in range(total)])
        ax.set_title(title)
        ax.grid(True, axis='x', alpha=0.25)

    draw_layout(axes[0], zones_equal, "连续等距分区的模块实际安装排布图", "#9ecae1", "tab:blue")
    draw_layout(axes[1], zones_aligned, "严格模块对齐分区的模块实际安装排布图", "#fcbba1", "tab:red")

    axes[1].set_xlabel("距离 (mm)")
    plt.tight_layout()
    plt.show()


# =============================
# 绘图：柱状图
# =============================
def plot_metrics_bar(metrics_equal, metrics_aligned):
    methods = ["等距分区", "模块对齐分区"]
    E_fit = [metrics_equal["E_fit"], metrics_aligned["E_fit"]]
    E_sep = [metrics_equal["E_sep"], metrics_aligned["E_sep"]]
    Score = [metrics_equal["综合得分"], metrics_aligned["综合得分"]]

    x = np.arange(len(methods))
    width = 0.22

    plt.figure(figsize=(10, 6))
    plt.bar(x - width, E_fit, width=width, label="E_fit")
    plt.bar(x, E_sep, width=width, label="E_sep")
    plt.bar(x + width, Score, width=width, label="综合得分")

    plt.xticks(x, methods)
    plt.ylabel("指标值")
    plt.title("分区评价指标对比")
    plt.legend()
    plt.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    plt.show()


# =============================
# 绘图：雷达图
# =============================
def plot_metrics_radar(metrics_equal, metrics_aligned):
    labels = ["拟合性", "分离性", "安装匹配", "均衡性", "尺寸合规"]

    fit_equal = np.exp(-metrics_equal["E_fit"] / 1000.0)
    fit_aligned = np.exp(-metrics_aligned["E_fit"] / 1000.0)

    max_sep = max(metrics_equal["E_sep"], metrics_aligned["E_sep"], 1e-6)
    sep_equal = metrics_equal["E_sep"] / max_sep
    sep_aligned = metrics_aligned["E_sep"] / max_sep

    heater_equal = np.exp(-metrics_equal["安装长度误差"] / 10.0)
    heater_aligned = np.exp(-metrics_aligned["安装长度误差"] / 10.0)

    balance_equal = metrics_equal["均衡性"]
    balance_aligned = metrics_aligned["均衡性"]

    compliance_equal = metrics_equal["尺寸合规率"]
    compliance_aligned = metrics_aligned["尺寸合规率"]

    values_equal = [fit_equal, sep_equal, heater_equal, balance_equal, compliance_equal]
    values_aligned = [fit_aligned, sep_aligned, heater_aligned, balance_aligned, compliance_aligned]

    values_equal += values_equal[:1]
    values_aligned += values_aligned[:1]

    angles = np.linspace(0, 2 * np.pi, len(labels), endpoint=False).tolist()
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
    ax.plot(angles, values_equal, linewidth=2, label="等距分区")
    ax.fill(angles, values_equal, alpha=0.15)
    ax.plot(angles, values_aligned, linewidth=2, label="模块对齐分区")
    ax.fill(angles, values_aligned, alpha=0.15)

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels)
    ax.set_ylim(0, 1.05)
    ax.set_title("分区质量雷达图对比", pad=20)
    ax.legend(loc="upper right", bbox_to_anchor=(1.25, 1.1))
    plt.tight_layout()
    plt.show()


# =============================
# 提取每个分区左中右三个固定物理位置点
# =============================
def extract_three_representative_points(zones, dist, temp, method_name):
    rows = []

    for z in zones:
        zone_id = z["id"]
        x0, x1 = z["range"]
        L = x1 - x0

        x_left = x0 + 0.30 * L
        x_mid = x0 + 0.50 * L
        x_right = x1 - 0.30 * L

        point_dict = {
            "左点": x_left,
            "中点": x_mid,
            "右点": x_right
        }

        for point_name, x in point_dict.items():
            t = np.interp(x, dist, temp)

            rows.append({
                "方法": method_name,
                "分区编号": zone_id,
                "点位名称": point_name,
                "坐标_mm": x,
                "温度_℃": t,
                "分区起点_mm": x0,
                "分区终点_mm": x1,
                "分区长度_mm": L
            })

    return pd.DataFrame(rows)


# =============================
# 导出辅助
# =============================
def zones_to_dataframe(zones, method_name, dist, temp):
    rows = []
    for z in zones:
        start_mm = z["range"][0]
        end_mm = z["range"][1]
        midpoint_mm = 0.5 * (start_mm + end_mm)
        midpoint_temp = np.interp(midpoint_mm, dist, temp)

        positions = compute_module_positions_in_zone(
            start_mm, end_mm, z["modules"],
            is_first=z["is_first"],
            is_last=z["is_last"],
            module_length=MODULE_LENGTH,
            module_gap=MODULE_GAP
        )

        module_pos_text = "; ".join([f"M{i+1}:[{p[0]:.1f},{p[1]:.1f}]" for i, p in enumerate(positions)])

        rows.append({
            "方法": method_name,
            "分区编号": z["id"],
            "起点_mm": start_mm,
            "终点_mm": end_mm,
            "中点_mm": midpoint_mm,
            "中点温度_℃": midpoint_temp,
            "分区长度_mm": z["size"],
            "模块数": z["modules"],
            "单模块长度_mm": MODULE_LENGTH,
            "模块间距_mm": MODULE_GAP,
            "模块本体总长度_mm": z["body_length"],
            "总间距长度_mm": z["gap_length"],
            "安装占位长度_mm": z["install_length"],
            "左侧外超_mm": z["left_overhang"],
            "右侧外超_mm": z["right_overhang"],
            "模块实际位置": module_pos_text,
            "分区平均温度_℃": z["avg_temp"]
        })
    return pd.DataFrame(rows)


def metrics_to_dataframe(metrics_equal, metrics_aligned):
    return pd.DataFrame([
        {
            "方法": "等距分区",
            "E_fit": metrics_equal["E_fit"],
            "E_sep": metrics_equal["E_sep"],
            "尺寸合规率": metrics_equal["尺寸合规率"],
            "安装长度误差": metrics_equal["安装长度误差"],
            "均衡性": metrics_equal["均衡性"],
            "内部违规数": metrics_equal["内部违规数"],
            "综合得分": metrics_equal["综合得分"],
            "总模块数": metrics_equal["总模块数"],
            "总安装占位长度": metrics_equal["总安装占位长度"],
            "总模块本体长度": metrics_equal["总模块本体长度"],
            "总间距长度": metrics_equal["总间距长度"],
            "左侧总外超长度": metrics_equal["左侧总外超长度"],
            "右侧总外超长度": metrics_equal["右侧总外超长度"]
        },
        {
            "方法": "模块对齐分区",
            "E_fit": metrics_aligned["E_fit"],
            "E_sep": metrics_aligned["E_sep"],
            "尺寸合规率": metrics_aligned["尺寸合规率"],
            "安装长度误差": metrics_aligned["安装长度误差"],
            "均衡性": metrics_aligned["均衡性"],
            "内部违规数": metrics_aligned["内部违规数"],
            "综合得分": metrics_aligned["综合得分"],
            "总模块数": metrics_aligned["总模块数"],
            "总安装占位长度": metrics_aligned["总安装占位长度"],
            "总模块本体长度": metrics_aligned["总模块本体长度"],
            "总间距长度": metrics_aligned["总间距长度"],
            "左侧总外超长度": metrics_aligned["左侧总外超长度"],
            "右侧总外超长度": metrics_aligned["右侧总外超长度"]
        }
    ])


def raw_data_to_dataframe(dist, temp):
    return pd.DataFrame({
        "距离_mm": dist,
        "温度_℃": temp
    })


# =============================
# Excel 美化
# =============================
def beautify_excel(filename):
    wb = load_workbook(filename)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    metric_fill = PatternFill("solid", fgColor="FFF2CC")
    point_fill = PatternFill("solid", fgColor="E2F0D9")
    header_font = Font(bold=True, color="000000")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF")
    )

    point_sheets = ["等距分区三点", "模块对齐分区三点"]

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    if ws.title in ["原始数据", "等距分区结果", "模块对齐分区结果", "等距分区三点", "模块对齐分区三点"]:
                        cell.number_format = "0.000"
                    elif ws.title == "评价指标":
                        cell.number_format = "0.0000"

        if ws.title == "评价指标":
            for row in ws.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    cell.fill = metric_fill
                    cell.font = Font(bold=True)
                    cell.alignment = center_align
                    cell.border = thin_border

        if ws.title in point_sheets:
            for row in ws.iter_rows(min_row=2, max_col=3):
                for cell in row:
                    cell.fill = point_fill
                    cell.alignment = center_align
                    cell.border = thin_border

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    cell_value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(cell_value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 45)

    wb.save(filename)


# =============================
# 导出 Excel
# =============================
def export_to_excel(filename, dist, temp, zones_equal, zones_aligned, metrics_equal, metrics_aligned):
    df_raw = raw_data_to_dataframe(dist, temp)
    df_equal = zones_to_dataframe(zones_equal, "等距分区", dist, temp)
    df_aligned = zones_to_dataframe(zones_aligned, "模块对齐分区", dist, temp)
    df_metrics = metrics_to_dataframe(metrics_equal, metrics_aligned)

    df_equal_3points = extract_three_representative_points(zones_equal, dist, temp, "等距分区")
    df_aligned_3points = extract_three_representative_points(zones_aligned, dist, temp, "模块对齐分区")

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df_raw.to_excel(writer, sheet_name="原始数据", index=False)
        df_equal.to_excel(writer, sheet_name="等距分区结果", index=False)
        df_aligned.to_excel(writer, sheet_name="模块对齐分区结果", index=False)
        df_equal_3points.to_excel(writer, sheet_name="等距分区三点", index=False)
        df_aligned_3points.to_excel(writer, sheet_name="模块对齐分区三点", index=False)
        df_metrics.to_excel(writer, sheet_name="评价指标", index=False)

    beautify_excel(filename)
    print(f"\nExcel 文件已导出并美化：{filename}")


# =============================
# 主程序
# =============================
if __name__ == "__main__":
    print("==============================================")
    print("分区程序启动：内部不允许超出，两端外边缘允许超出")
    print(f"模块本体长度: {MODULE_LENGTH:.1f} mm")
    print(f"模块间距: {MODULE_GAP:.1f} mm")
    print(f"模块安装节距: {MODULE_PITCH:.1f} mm")
    print(f"左右外边缘允许超出: {OUTER_EDGE_ALLOW:.1f} mm")
    print("==============================================")

    weights = compute_weights(distance, temperature, alpha=ALPHA)

    # 连续等距分区
    zones_equal_raw = equal_partition_continuous(TOTAL_LENGTH, EQUAL_K)
    zones_equal = build_zones_equal_continuous(distance, temperature, zones_equal_raw)
    metrics_equal = evaluate_zoning_quality(distance, temperature, zones_equal)

    # 严格模块对齐分区
    edges, zones_idx_aligned, mean_table = optimal_partition_aligned(distance, temperature, weights)
    zones_aligned = build_zones_aligned(edges, zones_idx_aligned, mean_table)
    metrics_aligned = evaluate_zoning_quality(distance, temperature, zones_aligned)

    print_metrics("连续等距分区评价", metrics_equal)
    print_metrics("严格模块对齐分区评价", metrics_aligned)

    # 图1：温度曲线 + 分区 + 模块真实排布
    plot_compare_equal_vs_aligned(
        distance, temperature,
        zones_equal, zones_aligned, edges,
        metrics_equal, metrics_aligned
    )

    # 图2：独立安装排布图
    plot_module_layout(zones_equal, zones_aligned)

    # 图3/4：评价图
    plot_metrics_bar(metrics_equal, metrics_aligned)
    plot_metrics_radar(metrics_equal, metrics_aligned)

    # 导出 Excel
    export_to_excel(
        filename=EXPORT_FILENAME,
        dist=distance,
        temp=temperature,
        zones_equal=zones_equal,
        zones_aligned=zones_aligned,
        metrics_equal=metrics_equal,
        metrics_aligned=metrics_aligned
    )