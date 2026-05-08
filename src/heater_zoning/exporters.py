from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .fonts import configure_matplotlib_fonts
from .models import AnalysisResult
from .reporting import build_report_frames, build_summary_cards, build_zone_summary_table
from .visualization import (
    build_metrics_bar_matplotlib,
    build_metrics_radar_matplotlib,
    build_module_layout_matplotlib,
    build_temperature_comparison_matplotlib,
)


def _sheet_title_match(sheet_title: str, *candidates: str) -> bool:
    return sheet_title in candidates


def _apply_metrics_conditional_formatting(sheet):
    green_fill = PatternFill("solid", fgColor="E8F5E9")
    red_fill = PatternFill("solid", fgColor="FDECEC")
    blue_fill = PatternFill("solid", fgColor="EAF2FF")

    header_map = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}
    equal_col = get_column_letter(header_map["等距分区"])
    aligned_col = get_column_letter(header_map["模块对齐分区"])
    direction_col = get_column_letter(header_map["指标方向"])
    score_col = get_column_letter(header_map["纳入综合得分"])
    better_col = get_column_letter(header_map["更优方案"])

    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        if sheet[f"{score_col}{row}"].value == "是":
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).fill = blue_fill

    equal_range = f"{equal_col}2:{equal_col}{max_row}"
    aligned_range = f"{aligned_col}2:{aligned_col}{max_row}"
    better_range = f"{better_col}2:{better_col}{max_row}"

    sheet.conditional_formatting.add(
        equal_range,
        FormulaRule(
            formula=[f'AND(${direction_col}2="越大越好",{equal_col}2>{aligned_col}2)', f'AND(${direction_col}2="越小越好",{equal_col}2<{aligned_col}2)'],
            fill=green_fill,
        ),
    )
    sheet.conditional_formatting.add(
        aligned_range,
        FormulaRule(
            formula=[f'AND(${direction_col}2="越大越好",{aligned_col}2>{equal_col}2)', f'AND(${direction_col}2="越小越好",{aligned_col}2<{equal_col}2)'],
            fill=green_fill,
        ),
    )
    sheet.conditional_formatting.add(
        better_range,
        FormulaRule(formula=[f'${better_col}2="等距分区"', f'${better_col}2="模块对齐分区"'], fill=green_fill),
    )
    sheet.conditional_formatting.add(
        better_range,
        FormulaRule(formula=[f'${better_col}2="按项目判断"'], fill=red_fill),
    )


def _apply_difference_conditional_formatting(sheet):
    green_fill = PatternFill("solid", fgColor="E8F5E9")
    amber_fill = PatternFill("solid", fgColor="FFF4E5")
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}
    better_col = get_column_letter(header_map["更优方案"])
    diff_col = get_column_letter(header_map["相对变化"])
    max_row = sheet.max_row

    sheet.conditional_formatting.add(
        f"{better_col}2:{better_col}{max_row}",
        FormulaRule(formula=[f'${better_col}2="等距分区"', f'${better_col}2="模块对齐分区"'], fill=green_fill),
    )
    sheet.conditional_formatting.add(
        f"{diff_col}2:{diff_col}{max_row}",
        FormulaRule(formula=[f"ABS({diff_col}2)>=0.1"], fill=amber_fill),
    )


def beautify_excel(workbook_path: Path):
    workbook = load_workbook(workbook_path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    metric_fill = PatternFill("solid", fgColor="FFF2CC")
    point_fill = PatternFill("solid", fgColor="E2F0D9")
    summary_fill = PatternFill("solid", fgColor="EEF6FF")
    header_font = Font(bold=True, color="000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

        if _sheet_title_match(sheet.title, "评价指标"):
            for row in sheet.iter_rows(min_row=2, max_col=4):
                for cell in row:
                    cell.fill = metric_fill
            _apply_metrics_conditional_formatting(sheet)

        if _sheet_title_match(sheet.title, "方案差异摘要"):
            for row in sheet.iter_rows(min_row=2, max_col=3):
                for cell in row:
                    cell.fill = summary_fill
            _apply_difference_conditional_formatting(sheet)

        if "三点" in sheet.title:
            for row in sheet.iter_rows(min_row=2, max_col=3):
                for cell in row:
                    cell.fill = point_fill

        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in column_cells:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            sheet.column_dimensions[col_letter].width = min(max_length + 4, 45)

    workbook.save(workbook_path)


def export_analysis_excel(result: AnalysisResult, output_path: Path) -> Path:
    frames = build_report_frames(result)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        frames.profile.to_excel(writer, sheet_name="原始数据", index=False)
        frames.equal_zones.to_excel(writer, sheet_name="等距分区结果", index=False)
        frames.aligned_zones.to_excel(writer, sheet_name="模块对齐分区结果", index=False)
        frames.equal_points.to_excel(writer, sheet_name="等距分区三点", index=False)
        frames.aligned_points.to_excel(writer, sheet_name="模块对齐三点", index=False)
        frames.metrics.to_excel(writer, sheet_name="评价指标", index=False)
        frames.differences.to_excel(writer, sheet_name="方案差异摘要", index=False)

    beautify_excel(output_path)
    return output_path


def _table_figure(dataframe: pd.DataFrame, title: str):
    rows = min(len(dataframe), 16)
    height = max(3.8, 0.45 * rows + 1.6)
    fig, axis = plt.subplots(figsize=(11.69, height))
    fig.patch.set_facecolor("white")
    axis.axis("off")
    axis.set_title(title, loc="left", fontsize=14, fontweight="bold", pad=12)
    table = axis.table(
        cellText=dataframe.head(16).values,
        colLabels=list(dataframe.columns),
        loc="center",
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1, 1.5)
    for (row, col), cell in table.get_celld().items():
        if row == 0:
            cell.set_facecolor("#d9eaf7")
            cell.set_text_props(weight="bold")
        cell.set_edgecolor("#cbd5e1")
    fig.tight_layout()
    return fig


def _summary_figure(result: AnalysisResult):
    cards = build_summary_cards(result)
    summary_df = build_zone_summary_table(result)
    fig, axes = plt.subplots(2, 1, figsize=(11.69, 8.27), gridspec_kw={"height_ratios": [1.2, 3.0]})
    fig.patch.set_facecolor("white")

    axes[0].axis("off")
    axes[0].text(0.0, 1.0, "Heater Zoning Optimizer 摘要", fontsize=18, fontweight="bold", va="top")
    y = 0.72
    for card in cards:
        axes[0].text(0.0, y, f"{card['label']}: {card['value']}", fontsize=12, va="top")
        y -= 0.15

    axes[1].axis("off")
    axes[1].set_title("分区汇总", loc="left", fontsize=13, fontweight="bold", pad=10)
    table = axes[1].table(
        cellText=summary_df.values,
        colLabels=list(summary_df.columns),
        loc="center",
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1, 1.45)
    for (row, col), cell in table.get_celld().items():
        if row == 0:
            cell.set_facecolor("#d9eaf7")
            cell.set_text_props(weight="bold")
        cell.set_edgecolor("#cbd5e1")

    fig.tight_layout()
    return fig


def export_summary_pdf(result: AnalysisResult, output_path: Path) -> Path:
    frames = build_report_frames(result)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    figures = [
        _summary_figure(result),
        build_temperature_comparison_matplotlib(result.profile_df, result.equal_zones, result.aligned_zones),
        build_module_layout_matplotlib(result.equal_zones, result.aligned_zones),
        build_metrics_bar_matplotlib(result.equal_metrics, result.aligned_metrics),
        build_metrics_radar_matplotlib(result.equal_metrics, result.aligned_metrics),
        _table_figure(frames.metrics, "评价指标"),
        _table_figure(frames.differences, "方案差异摘要"),
        _table_figure(frames.equal_zones, "等距分区结果"),
        _table_figure(frames.aligned_zones, "模块对齐分区结果"),
    ]

    with PdfPages(output_path) as pdf:
        for figure in figures:
            pdf.savefig(figure, bbox_inches="tight")
            plt.close(figure)

    return output_path


configure_matplotlib_fonts()
