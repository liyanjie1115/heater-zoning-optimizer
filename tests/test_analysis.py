import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.heater_zoning import AnalysisConfig, analyze_profile, sample_profile_dataframe
from src.heater_zoning.cli import main as cli_main
from src.heater_zoning.exporters import export_analysis_excel, export_summary_pdf
from src.heater_zoning.io_utils import normalize_profile_dataframe
from src.heater_zoning.reporting import build_analysis_notes, build_config_snapshot, build_recommendation_summary
from src.heater_zoning.runflow import run_analysis_pipeline
from src.heater_zoning.settings import AppSettings


class AnalysisSmokeTest(unittest.TestCase):
    def test_sample_profile_analysis_runs(self):
        result = analyze_profile(sample_profile_dataframe(), AnalysisConfig())
        self.assertEqual(len(result.equal_zones), 8)
        self.assertGreaterEqual(len(result.aligned_zones), 1)
        self.assertLessEqual(len(result.aligned_zones), 8)
        self.assertGreater(result.equal_metrics.total_modules, 0)
        self.assertGreater(result.aligned_metrics.total_modules, 0)
        self.assertGreater(result.aligned_metrics.composite_score, result.equal_metrics.composite_score)
        self.assertGreater(result.aligned_metrics.gradient_capture_score, result.equal_metrics.gradient_capture_score)

    def test_normalize_profile_dataframe_accepts_alias_columns(self):
        raw = pd.DataFrame({"距离": [500.0, 0.0, 250.0, 250.0], "温度": [320.0, 650.0, 450.0, 451.0]})
        normalized = normalize_profile_dataframe(raw)
        self.assertEqual(list(normalized.columns), ["distance_mm", "temperature_c"])
        self.assertEqual(normalized["distance_mm"].tolist(), [0.0, 250.0, 500.0])

    def test_export_analysis_excel_creates_expected_sheets(self):
        result = analyze_profile(sample_profile_dataframe(), AnalysisConfig())
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "report.xlsx"
            export_analysis_excel(result, output_path)
            self.assertTrue(output_path.exists())
            workbook = load_workbook(output_path)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "原始数据",
                    "等距分区结果",
                    "模块对齐分区结果",
                    "等距分区三点",
                    "模块对齐三点",
                    "评价指标",
                    "方案差异摘要",
                    "论文变量映射_分区",
                    "Fig3_温度边界数据",
                    "Fig3_模块排布数据",
                    "Table4_分区策略对比",
                    "分区明细_论文命名",
                ],
            )
            self.assertIn("T_tar(s)", [cell.value for cell in workbook["论文变量映射_分区"]["A"]])
            self.assertIn("E_part", [cell.value for cell in workbook["Table4_分区策略对比"][1]])

    def test_cli_generates_report(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            exit_code = cli_main(["--output-dir", tmpdir, "--output-name", "cli_report.xlsx", "--json"])
            self.assertEqual(exit_code, 0)
            self.assertTrue((Path(tmpdir) / "cli_report.xlsx").exists())

    def test_cli_accepts_preset(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            exit_code = cli_main(["--output-dir", tmpdir, "--output-name", "preset_report.xlsx", "--preset", "compact", "--json"])
            self.assertEqual(exit_code, 0)
            self.assertTrue((Path(tmpdir) / "preset_report.xlsx").exists())

    def test_run_analysis_pipeline_uses_sample_source(self):
        artifacts = run_analysis_pipeline(config=AnalysisConfig(), source="sample")
        self.assertTrue(artifacts.export_path.exists())
        self.assertEqual(artifacts.summary_cards[0]["label"], "推荐方案")

    def test_metrics_dataframe_has_user_facing_columns(self):
        artifacts = run_analysis_pipeline(config=AnalysisConfig(), source="sample")
        metrics = artifacts.frames.metrics
        self.assertEqual(
            list(metrics.columns),
            ["指标", "指标方向", "纳入综合得分", "更优方案", "等距分区", "模块对齐分区", "指标说明"],
        )
        self.assertIn("综合得分", metrics["指标"].tolist())

    def test_difference_dataframe_exists(self):
        artifacts = run_analysis_pipeline(config=AnalysisConfig(), source="sample")
        differences = artifacts.frames.differences
        self.assertEqual(
            list(differences.columns),
            ["指标", "指标方向", "等距分区", "模块对齐分区", "绝对差值(模块对齐-等距)", "相对变化", "更优方案"],
        )
        self.assertIn("综合得分", differences["指标"].tolist())

    def test_partition_paper_frames_exist(self):
        artifacts = run_analysis_pipeline(config=AnalysisConfig(), source="sample")
        frames = artifacts.frames
        self.assertEqual(
            list(frames.table4_partition_comparison.columns),
            [
                "Method",
                "方法",
                "K",
                "E_part",
                "reduction_%",
                "compliance_%",
                "total_modules",
                "total_install_length_mm",
                "heater_mismatch_mm",
            ],
        )
        self.assertIn("T_tar(s)", frames.partition_variable_mapping["paper_symbol"].tolist())
        self.assertIn("zone_interval", frames.fig3_temperature_boundaries["row_type"].tolist())
        self.assertIn("module_start_mm", frames.fig3_module_layout.columns)
        self.assertIn("Omega_k", frames.paper_partition_details.columns)

    def test_recommendation_helpers_return_content(self):
        result = analyze_profile(sample_profile_dataframe(), AnalysisConfig())
        recommendation = build_recommendation_summary(result)
        snapshot = build_config_snapshot(result)
        notes = build_analysis_notes(result)
        self.assertIn("winner", recommendation)
        self.assertGreater(len(recommendation["reasons"]), 0)
        self.assertGreater(len(snapshot), 0)
        self.assertGreater(len(notes), 0)

    def test_export_summary_pdf_creates_file(self):
        result = analyze_profile(sample_profile_dataframe(), AnalysisConfig())
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "summary.pdf"
            export_summary_pdf(result, output_path)
            self.assertTrue(output_path.exists())
            self.assertGreater(output_path.stat().st_size, 0)

    def test_app_settings_round_trip(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            settings_path = Path(tmpdir) / "settings.json"
            settings = AppSettings()
            settings.add_recent_file("D:/data/profile.csv")
            settings.save_template("default", AnalysisConfig())
            settings.save(settings_path)

            restored = AppSettings.load(settings_path)
            self.assertEqual(restored.recent_files[0], "D:/data/profile.csv")
            self.assertIn("default", restored.templates)
            self.assertEqual(restored.load_template("default").module_gap, 10.0)


if __name__ == "__main__":
    unittest.main()
